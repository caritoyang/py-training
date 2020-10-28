import openpyxl

# Load Workbook - create object
wk = openpyxl.load_workbook("C:/Users/carol/Documents/XLRD1.xlsx")

# read sheet names
print(wk.sheetnames)
print("Active Sheet is " + wk.active.title)

# Create Object of a sheet on which you want to work
sh = wk['Sheet1']
print(sh.title)

# print a cell value
print(sh['C3'].value)
print(sh['D1'].value)

# create cell object (row, col - not index)
c1 = sh.cell(1,1)
print(c1.value)

print(c1.row)
print(c1.column)

# READ COMPLETE TABLE
# Find rows having data
rows = sh.max_row
columns = sh.max_column

print("Total Rows are: " + str(rows))
print("Total Columns are: " + str(columns))

for i in range (1,rows+1):
    for j in range (1, columns+1):
        c = sh.cell(i,j)
        print(c.value)

for r in sh['A1':'D3']:
    for c in r:
        print(c.value)